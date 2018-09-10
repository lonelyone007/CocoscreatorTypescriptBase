# -*- coding: utf-8 -*-
# cunkai
import errno
import os
import shutil
import re
import csv
import platform
import sys
import codecs
import copy
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding("utf-8")

sysType = platform.system()
ESC = chr(27)
# 字体
a_default = 0
a_bold = 1
a_italic = 3
a_underline = 4
# 前景颜色
fg_black = 30
fg_red = 31
fg_green = 32
fg_yellow = 33
fg_blue = 34
fg_magenta = 35
fg_cyan = 36
fg_white = 37
# 后景颜色
bg_black = 40
bg_red = 41
bg_green = 42
bg_yellow = 43
bg_blue = 44
bg_magenta = 45
bg_cyan = 46
bg_white = 47


def color_code(a):
    return ESC + "[%dm" % a


def color_str(s, *args):
    if (sysType == "Windows"):
        return s
    cs = ""
    for a in args:
        cs += color_code(a)
    cs += s
    cs += color_code(a_default)
    return cs


# currentPath = os.path.split(os.path.realpath(__file__))[0]
currentPath = os.path.dirname(os.path.abspath(__file__))
projectRootPath = os.path.abspath(os.path.join(currentPath, ".."))

csvPath = os.path.join(currentPath, "csv_folder/")
xlsx_path = os.path.join(csvPath, "数值表.xlsx")

csharp = os.path.join(currentPath, "csharp_folder/")
csvDataParentName = "CSVReader"
dataManageName = "dataManage"

defaultcs = os.path.join(currentPath, "default/defaultclass.ts")


defaultConfigManage = os.path.join(
    projectRootPath, "Excelplus/default/defaultmanager.ts")
configManagerPath = os.path.join(
    projectRootPath, "assets/Script/Config/ConfigManager.ts")

defaultCsvReader = os.path.join(
    projectRootPath, "Excelplus/default/CSVReader.ts")
csvReaderPath = os.path.join(
    projectRootPath, "assets/Script/Config/CSVReader.ts")

defaultConfigPath = os.path.join(currentPath, "default/Config.ts")
targetConfigPath = os.path.join(
    projectRootPath, "assets/Script/Config/Config.ts")
# defaultCsvDataParent = path + "default/CSVReader.ts"


# 存放项目位置
mylocation = os.path.join(projectRootPath, "assets/Script/Config/CsvClass/")
# 转换后文件路径
reslocation = os.path.join(projectRootPath, "assets/configs/")

print(color_str("SYS: " + sysType, bg_blue, bg_cyan, a_bold))
print "coding: " + sys.getdefaultencoding()

classList = []
# CSV文件所在路径
csvFI = ""
# cs文件所在路径
csFI = ""


def toCsString(line):
    isItem = 0

    data = '\"'

    for i in range(len(line)):
        item = line[i]
        if item != '':
            isItem = 1
            data += item
            if i < len(line) - 1:
                data += '|'
        else:
            data += '-1'
            if i < len(line) - 1:
                data += '|'

    data += '\"'

    if isItem == 0:
        return ''
    return data


def openFile(path, ftype):
    fp = None
    mkdir(os.path.dirname(path))

    if (sysType == "Windows"):
        try:
            fp = open(path, ftype, encoding='utf-8')
        except:
            fp = open(path, ftype)
    else:
        fp = open(path, ftype)
    return fp


def mkdir(path):
    path = path.strip()
    path = path.rstrip("\\")
    isExists = os.path.exists(path)

    if not isExists:
        # print path+' OK'
        os.makedirs(path)
        return True
    else:
        # print path+'OK'
        return False


def toIfFromClass(name):
    code = "\t\t\tif(className==\"" + name + \
        "\")\n\t\t\t{" + "\n\t\t\t\t" + "return " + \
        name + ".Instance;" + "\n\t\t\t}"
    return code

# 给文件加按键字段,以及说明

def copy_create_dir(src, dest) :    
    try:
        shutil.copy(src, dest)
    except IOError as e:
        # ENOENT(2): file does not exist, raised also on missing dest parent dir
        if e.errno != errno.ENOENT:
            raise
        # try creating parent directories
        print dest
        os.makedirs(os.path.dirname(dest))
        shutil.copy(src, dest)


def CreateTheKeyValue(cavfilePath, csfilePath):
    # str-----------------------------------------------------
    # 添加的替换字符串
    # 查询键
    keyText = ""
    # 标题
    nameText = ""
    # 实体类列
    classText = ""
    # 实体类数据初始化列
    classInitText = ""

    fileName = os.path.splitext(os.path.basename(csfilePath))[0]
    with open(cavfilePath, 'rb') as f:
        # 读取文件
        reader = csv.reader(f)
        # 字节说明
        explainList = []
        # 类型说明
        typeList = []
        # 名字列表
        nameList = []
        i = 0
        # 读取每一行
        for row in reader:
            i += 1
            if i == 1:
                # 读取表的说明
                s1 = row[0].decode('utf-8')
                nameText = "/** " + s1 + " */"
                row[0] = s1
                # 读取键值的说明
                explainList = row
            if i == 2:
                typeList = row  # 变量的类型
                funclist = copy.deepcopy(row)  # 取得变量的函数类型
            if i == 3:
                nameList = row
        # break
        for i in range(len(typeList)):
            if nameList[i] == 'x' or nameList[i] == '':
                continue

            nameCell = nameList[i]
            shuomin = explainList[i]
            if typeList[i].rstrip() == "array":
                typeList[i] = "string"
                funclist[i] = "string"
            elif typeList[i].rstrip() == "enumflag":
                funclist[i] = "int"
            if typeList[i].rstrip() == "enumflag" or typeList[i].rstrip() == "int" or typeList[i].rstrip() == "float":
                typeList[i] = "number"
            if i > 0:
                pass
                # shuomin = shuomin.decode('gbk')
            keyText += "\t/** " + shuomin + " */\n\tpublic static _" + \
                nameCell + "=\"" + nameCell + "\";\n"

            classText += "\t/** " + shuomin + " */\n\tpublic " + \
                         " " + nameCell + ":" + \
                         typeList[i] + \
                         ";\n"
            classInitText += "\t\t\t_line." + nameCell + " = " + fileName + \
                ".Instance.get" + funclist[i] + \
                "ById(id, " + "this._" + nameCell + ");\n"

    # 标识符字符串
    # TheReservedText,ReservedTextName

    # 看一下这个文件是否存在
    if not os.path.exists(csfilePath):
        print "文件不存在".decode('utf-8'), csfilePath
        raw_input(unicode('按回车键退出...', 'utf-8').encode('gbk'))
        exit(-1)  # ，不存在就退出
    else:
        lines = open(csfilePath).readlines()  # 打开文件，读入每一行
        fp = open(csfilePath, 'w')  # 打开你要写得文件pp2.txt
        for s in lines:
            fp.write(s.replace('//TheReservedText', keyText).
                     replace('//ReservedTextName', nameText).
                     replace('//TheReservedValueText', classText).
                     replace('//TheReservedInitializeText', classInitText))  # replace是替换，write是写入
        fp.close()  # 关闭文件
        print csfilePath, "加注释OK".decode('utf-8')

# end-----------------------------------------------------


def process_configmanater(files):
    contentFormat = "\t\t\tConfigManager.csvCache[\"{fileName}\"] = new CSVReader(\"{fileName}\").LoadSync(config.{fileName});\n"
    copy_create_dir(defaultConfigManage, configManagerPath)
    copy_create_dir(defaultCsvReader, csvReaderPath)
    if (not os.path.exists(configManagerPath)):
        print "文件不存在".decode('utf-8'), configManagerPath
        raw_input(unicode('按回车键退出...', 'utf-8').encode('gbk'))
        exit(-1)  # ，不存在就退出
    else:
        lines = open(configManagerPath).readlines()  # 打开文件，读入每一行
        fp = open(configManagerPath, 'w')  # 打开你要写得文件pp2.txt

        content = ""
        for f in files:
            if not str.endswith(f, '.csv'):
                continue
            fileName = os.path.splitext(f)[0]
            content += contentFormat.format(fileName=fileName)
        for s in lines:
            fp.write(s.replace('//allcsvFile', content))  # replace是替换，write是写入
        fp.close()  # 关闭文件
        print "ConfigManager更新成功".decode('utf-8')

    contentFormat = "\t@property(cc.TextAsset)\n\t{fileName}: cc.TextAsset = null;\n"
    copy_create_dir(defaultConfigPath, targetConfigPath)
    if (not os.path.exists(defaultConfigPath)):
        print "文件不存在".decode('utf-8'), defaultConfigPath
        raw_input(unicode('按回车键退出...', 'utf-8').encode('gbk'))
        exit(-1)  # ，不存在就退出
    else:
        lines = open(targetConfigPath).readlines()  # 打开文件，读入每一行
        fp = open(targetConfigPath, 'w')  # 打开你要写得文件pp2.txt

        content = ""
        for f in files:
            if not str.endswith(f, '.csv'):
                continue
            fileName = os.path.splitext(f)[0]
            content += contentFormat.format(fileName=fileName)
        for s in lines:
            fp.write(s.replace('//Content', content))  # replace是替换，write是写入

        fp.close()  # 关闭文件
        print "Config.ts更新成功".decode('utf-8')


def check_contain_chinese(check_str):
    for ch in check_str.decode('utf-8'):
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False

letter = [ 
    'A', 
    'B', 
    'C', 
    'D', 
    'E', 
    'F', 
    'G', 
    'H', 
    'I', 
    'J', 
    'K', 
    'L', 
    'M', 
    'N', 
    'O', 
    'P', 
    'Q', 
    'R', 
    'S', 
    'T', 
    'U', 
    'V', 
    'W', 
    'X', 
    'Y', 
    'Z', 

    'AA', 
    'AB', 
    'AC', 
    'AD', 
    'AE', 
    'AF', 
    'AG', 
    'AH', 
    'AI', 
    'AJ', 
    'AK', 
    'AL', 
    'AM', 
    'AN', 
    'AO', 
    'AP', 
    'AQ', 
    'AR', 
    'AS', 
    'AT', 
    'AU', 
    'AV', 
    'AW', 
    'AX', 
    'AY', 
    'AZ', 

    'BA', 
    'BB', 
    'BC', 
    'BD', 
    'BE', 
    'BF', 
    'BG', 
    'BH', 
    'BI', 
    'BJ', 
    'BK', 
    'BL', 
    'BM', 
    'BN', 
    'BO', 
    'BP', 
    'BQ', 
    'BR', 
    'BS', 
    'BT', 
    'BU', 
    'BV', 
    'BW', 
    'BX', 
    'BY', 
    'BZ', 
]

def convert_xlsx_to_csv(filepath):
    print filepath
    filename = filepath.decode("utf-8")
    wb = load_workbook(filename, data_only=True)
    #  -*- coding:gbk -*-

    # 循环遍历所有sheet
    for sheet in wb.sheetnames:
        if check_contain_chinese(sheet):
            continue
        # print('\n\n第'+str(i+1)+'个sheet: ' + sheet.title+'->>>')

        csv_filename = csvPath + '{sheet}.csv'.format(
            sheet=sheet
        )

        csv_file = openFile(csv_filename, 'wb')
        csv_file_writer = csv.writer(csv_file)

        sheet_ranges = wb[sheet]
        srl = list(sheet_ranges.columns)
        for i in range(len(srl) - 1, 0, -1):
            if srl[i-1][2].value == srl[i][2].value and not srl[i][2].value == None: 
                print '合并列', srl[i-1][2].value
                srl[i][1].value = ''
                srl[i][2].value = ''
                for j in range(len(list(srl[i]))):
                    if j !=2 and j != 1:
                        srl[i-1][j].value = '{0}|{1}'.format(srl[i-1][j].value, srl[i][j].value)
                        srl[i][j].value = ''
                    # mergeStr = '{0}{2}:{1}{2}'.format(letter[i-1], letter[i], j+1)

                    # sheet_ranges.merge_cells(mergeStr)

        for row in sheet_ranges.rows:
            row_container = []
            for cell in row:
                value = ""
                if type(cell.value) == unicode:
                    value = cell.value.encode('utf-8')
                else:
                    value = str(cell.value)

                value = value.replace(',', "comma").replace('None', "")

                row_container.append(value)
            csv_file_writer.writerow(row_container)
        csv_file.close()


def start(rootDir):

    list_dirs = os.walk(rootDir)

    # remove old csharp
    isExists = os.path.exists(csharp)
    if isExists:
        shutil.rmtree(csharp)

    print "-------------------------------------------------------"
    print "当前所在目录".decode('utf-8'), projectRootPath
    print "-----------------------------------"

    
    for root,_,files in os.walk(csvPath) :
        for f in files:
            if str.endswith(f, '.xlsx') and not str.startswith(f, '~'):
                convert_xlsx_to_csv(os.path.join(root,f))

    for root, _, files in list_dirs:
        process_configmanater(files)
        for f in files:
            if not str.endswith(f, '.csv'):
                continue
            filePath = os.path.join(root, f)
            fname, fextension = os.path.splitext(f)

            # 类文件存放地方
            csharpFilePath = mylocation + fname + ".ts"

            csvFI = filePath
            csFI = csharpFilePath

            # csharpFilePath = csharpFilePath.replace(csvPath,csharp)
            if fextension == ".csv":
                # print(f)
                mkdir(csharp + "data")
                # copy default.cs
                extendFile = currentPath + "/extends/" + fname + ".ts"
                if os.path.exists(extendFile):
                    copy_create_dir(extendFile, csharpFilePath)
                else:
                    copy_create_dir(defaultcs, csharpFilePath)

                # add classList
                classList.append(fname)

                # 配置文件utf-8拷贝//csv文件存放地方
                csvFolder = reslocation

                csvFilePath = csvFolder + f
                # ；compatible
                fp = None
                sf = open(filePath, 'rU')
                s = sf.read()
                fp = openFile(csvFilePath, 'w+')
                fp.write(s)
                fp.flush()
                sf.close()
                fp.close()

                fp = None

                fp = openFile(csvFilePath, 'r')
                alllines = fp.readlines()
                fp.close()
                fp = openFile(csvFilePath, 'w+')
                alllines = combine_rows(alllines)
                # write start

                for eachline in alllines:
                    # class name ,替换名字
                    a = re.sub(',', '#', eachline)
                    a = re.sub('array', 'string', a)
                    a = re.sub('enumflag', 'int', a)
                    # 逗号作为分隔符在配表示不允许使用, 所以用comma号代替
                    a = re.sub('comma', ',', a)
                    fp.writelines(a)
                fp.close()

                fp = None

                fp = openFile(csharpFilePath, 'r')
                alllines = fp.readlines()
                fp.close()
                fp = openFile(csharpFilePath, 'w+')

                # write start
                for eachline in alllines:
                    # class name
                    a = re.sub('defaultclass', fname, eachline)
                    # csvDataParent
                    a = re.sub('defaultCsvDataParent', csvDataParentName, a)
                    fp.writelines(a)
                fp.close()

            print "创建" + fname + "成功,开始添加键值序列".decode('utf-8')
            # print csvFI
            # print csFI
            CreateTheKeyValue(csvFI, csFI)


def combine_rows(lines):
    dicList = []  # 转换后的表格数组
    deleteIndexes = []
    typeList = []

    for i in range(len(lines)):
        line = lines[i]
        if i > 2:
            lineList = line.split(',')
            line = ConvertGrid(lineList, typeList)  # 转换需要转换的格式
            index = FindIndexInList(dicList, lineList[0])  # lineList[0] is id
            if index != -1:  # 合并数组列
                dicLine = dicList[index]
                dicList[index] = ComblineLine(dicLine, line, typeList)
            else:
                dicList.append(line)
        elif i == 1:
            dicList.append(line)
            typeList = line.strip('\n').split(',')
        elif i == 2:
            dicList.append(line)
            lineVars = line.strip('\n').split(',')
            deleteIndexes = FindIndexesInList(lineVars, ['x', ''])
        else:
            dicList.append(line)

    # 删除不需要的字段
    for i in range(len(dicList)):
        lineList = dicList[i].split(',')
        newLineList = []

        # 将需要导出的列放入新的数组中
        for j in range(len(lineList)):
            if j not in deleteIndexes:
                newLineList.append(lineList[j])

        # 如果是最后列,需要检查补齐换行符
        if not '\n' in newLineList[len(newLineList)-1]:
            newLineList[len(newLineList)-1] += '\n'
        dicList[i] = ','.join(map(str, newLineList))
    return dicList


def FindIndexInList(l, id):
    for i, item in enumerate(l):
        if item.split(',')[0] == id:
            return i
    return -1


def FindIndexesInList(l, ids):
    arr = []
    for i, item in enumerate(l):
        for id in ids:
            if item.split(',')[0] == id:
                arr.append(i)
    return arr


def ConvertGrid(lineList, typeList):
    for i in range(len(lineList)):
        if typeList[i].rstrip() == "enumflag":
            enumArr = lineList[i].split('|')
            num = 0
            for e in enumArr:
                if (int(e) > 0):
                    num |= (1 << (int(e)))
            lineList[i] = str(num)

            # 如果是最后一项,需要手动补齐换行符
            if i == len(lineList)-1:
                lineList[i] += '\n'
    return ','.join(map(str, lineList))


def ComblineLine(line1, line2, typeList):
    listLine1 = line1.split(',')
    listLine2 = line2.split(',')
    for i in range(len(listLine1)):
        if typeList[i].rstrip() == "array":
            listLine1[i] = listLine1[i].rstrip() + "!" + listLine2[i]

    if (sysType == "Windows"):
        listLine1.replace("\r", "")
    return ','.join(map(str, listLine1))


start(csvPath)

print(color_str("GOOD", fg_white, bg_green, a_italic))
# coding=utf-8

if (sysType == "Windows"):
    raw_input(unicode(
        '-------------------------------------------------------按回车键退出...', 'utf-8').encode('gbk'))
else:
    raw_input(unicode(
        '-------------------------------------------------------按回车键退出...', 'utf-8').encode('utf-8'))
