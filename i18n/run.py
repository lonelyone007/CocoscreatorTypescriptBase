# -*- coding: utf-8 -*-
# 将表格配置的转换成多语言文件
import os
import shutil
import re
import csv
import platform
import sys
import codecs
import copy
import itertools
from openpyxl import load_workbook

reload(sys)
sys.setdefaultencoding("utf-8")

sysType = platform.system()
ESC = chr(27)
# 字体
a_default = 0
a_bold = 1
a_italic = 3
a_underline = 4
# 前景颜色
fg_black = 30
fg_red = 31
fg_green = 32
fg_yellow = 33
fg_blue = 34
fg_magenta = 35
fg_cyan = 36
fg_white = 37
# 后景颜色
bg_black = 40
bg_red = 41
bg_green = 42
bg_yellow = 43
bg_blue = 44
bg_magenta = 45
bg_cyan = 46
bg_white = 47

current_path = os.path.split(os.path.realpath(__file__))[0]
root_path = os.path.abspath(os.path.join(current_path, ".."))

# 存放项目位置
target_path = root_path + "/assets/resources/i18n/"

xlsx_path = current_path + "/csv_folder/i18n.xlsx"
defaultjs = current_path + "/default/default.js"

default_interface = current_path + "/default/interface.ts"
target_interface = root_path + "/assets/Script/i18n/interface.ts"

def color_code(a):
    return ESC + "[%dm" % a

def color_str(s, *args):
    if (sysType == "Windows"):
        return s
    cs = ""
    for a in args:
        cs += color_code(a)
    cs += s
    cs += color_code(a_default)
    return cs

def openFile(name, ftype):
    fp = None
    if (sysType == "Windows"):
        try:
            fp = open(name, ftype, encoding='utf-8')
        except:
            fp = open(name, ftype)
    else:
        fp = open(name, ftype)

    return fp

def mkdir(path):
    path = path.strip()
    path = path.rstrip("\\")
    isExists = os.path.exists(path)
    if not isExists:
        # print path+' OK'
        os.makedirs(path)
        return True
    else:
        # print path+'OK'
        return False

def copy_file(src, target):
    if not os.path.exists(target):
        mkdir(target)
    shutil.copy(src, target)

# end-----------------------------------------------------

def convert_xlsx_to_js(filepath):
    filename = filepath.decode("utf-8")
    wb = load_workbook(filename, data_only=True)
    for sheet in wb.sheetnames:
        sheet_ranges = wb[sheet]
        cols = list(sheet_ranges.columns)
        ids = cols[0]
        zhs = cols[1]
        for col in cols[1:len(cols)]:
            output_language_to_file(col[2].value, ids, col)

        interface_content = ""
        i = 3
        for id in ids[3:len(ids)]:
            interface_content += "/** {3} */\n\tstatic get {0}() {1}return i18n.t('{0}'){2} \n    ".format(id.value, '{', '}', zhs[i].value)
            i += 1
        copy_file(default_interface, target_interface)
        # shutil.copy(default_interface, target_interface)
        operate_file(target_interface, lambda fp, line: fp.write(line.replace('//Contents', interface_content)))

def output_language_to_file(name, ids, contents):
    target = target_path + name + ".js"
    i = 3
    content = ""
    for id in ids[3:len(ids)]:
        content += "'{0}': '{1}',\n    ".format(id.value, contents[i].value)
        i += 1

    shutil.copy(defaultjs, target)
    operate_file(target, lambda fp, line: fp.write(line.replace('//Contents', content).
                                                  replace('//Language', name))
    )
    print target, "文件导出成功".decode('utf-8')

def operate_file(path, operate):
    lines = open(path).readlines()
    fp = open(path, 'w') 
    for line in lines:
        operate(fp, line)
    fp.close()

def start():
    projectRootPath = os.path.abspath(os.path.join(current_path, ".."))
    print  "-------------------------------------------------------"
    print  "当前所在目录".decode('utf-8'), projectRootPath
    print  "-----------------------------------"
    convert_xlsx_to_js(xlsx_path)

print(color_str("系统信息: " + sysType, bg_blue, bg_cyan, a_bold))
print "encoding: " + sys.getdefaultencoding()
start()

print(color_str("GOOD", fg_white, bg_green, a_italic))
# coding=utf-8

if (sysType == "Windows"):
    raw_input(unicode('-------------------------------------------------------按回车键退出...', 'utf-8').encode('gbk'))
else:
    raw_input(unicode('-------------------------------------------------------按回车键退出...', 'utf-8').encode('utf-8'))
